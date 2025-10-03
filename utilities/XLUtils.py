
import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def getColumnCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def readData(file, sheet_name, rownum, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=rownum, column=columno).value


def writeData(file, sheet_name, rownum, columno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=rownum, column=columno).value = data
    workbook.save(file)


def fillGreenColor(file, sheet_name, rownum, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    greenFill = PatternFill(start_color='FF60B212', end_color='FF60B212', fill_type="solid")
    sheet.cell(row=rownum, column=columno).fill = greenFill
    workbook.save(file)


def fillRedColor(file, sheet_name, rownum, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type="solid")
    sheet.cell(row=rownum, column=columno).fill = redFill
    workbook.save(file)






# import openpyxl
# from openpyxl.styles import PatternFill
#
#
# def getRowCount(file,sheetName):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     return sheet.max_row
#
# def getColumnCount(file,sheetName):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     return sheet.max_column
#
# def readData(file,sheetName,rownum,columno):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     return sheet.cell(rownum,columno).value
#
# def writeData(file,sheetName,rownum,columno,data):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     return sheet.cell(rownum,columno).value=data
#     workbook.save(file)
#
# def fillGreenColor(file,sheetName,rownum,columno):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     greenFill=PatternFill(start_color='60b212',end_color='60b212',fill_type="solid")
#     sheet.cell(rownum,columno).fill=greenFill
#     workbook.save(file)
#
# def fillRedColor(file,sheetName,rownum,columno):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     redFill=PatternFill(start_color='ff0000',end_color='ff0000',fill_type="solid")
#     sheet.cell(rownum,columno).fill=redFill
#     workbook.save(file)